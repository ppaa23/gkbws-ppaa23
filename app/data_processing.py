import os
import pandas as pd
import numpy as np
import warnings
import functools
from openpyxl import load_workbook
from app.logger import get_logger

logger = get_logger()


def get_data_file_path():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_dir, 'app', 'static', 'data', 'NIHMS1635539-supplement-1635539_Sup_tab_4.xlsx')


@functools.lru_cache(maxsize=1)
def load_volcano_data():
    """Load data for volcano plot from S4B limma results sheet with caching."""
    file_path = get_data_file_path()
    logger.info(f"Loading volcano data from {file_path} (first load or cache miss)")

    if not os.path.exists(file_path):
        logger.error(f"Excel file not found at {file_path}")
        raise FileNotFoundError(f"Excel file not found at {file_path}")

    # Identify the correct header row
    wb = load_workbook(filename=file_path, read_only=True)
    sheet = wb['S4B limma results']

    header_row = None
    for i in range(1, 10):
        if 'EntrezGeneSymbol' in [str(cell.value).strip() for cell in list(sheet.rows)[i - 1]]:
            header_row = i - 1
            break

    wb.close()

    if header_row is not None:
        df = pd.read_excel(file_path, sheet_name='S4B limma results', header=header_row)
        logger.info(f"Found header row at position {header_row}")
    else:
        logger.warning("Could not find 'EntrezGeneSymbol' in the first 10 rows. Falling back to header=2.")
        df = pd.read_excel(file_path, sheet_name='S4B limma results', header=2)

    # Validate expected columns
    if 'EntrezGeneSymbol' not in df.columns:
        logger.error(f"Column 'EntrezGeneSymbol' not found. Available columns: {df.columns.tolist()}")
        raise ValueError(f"Expected column 'EntrezGeneSymbol' not found. Available columns: {df.columns.tolist()}")

    # Handle column name variations
    if 'adj.P.Val' not in df.columns and 'P.Value' in df.columns:
        df.rename(columns={'P.Value': 'adj.P.Val'}, inplace=True)
        logger.info("Renamed 'P.Value' column to 'adj.P.Val'")

    if 'logFC' not in df.columns:
        fc_col = next((col for col in df.columns if 'FC' in col or 'fold' in col.lower()), None)
        if fc_col:
            df.rename(columns={fc_col: 'logFC'}, inplace=True)
            logger.info(f"Renamed '{fc_col}' column to 'logFC'")
        else:
            logger.error(f"No column related to fold change found. Available columns: {df.columns.tolist()}")
            raise ValueError(f"No column related to fold change found. Available columns: {df.columns.tolist()}")

    # Calculate -log10(p-value) for volcano plot
    min_pval = df['adj.P.Val'][df['adj.P.Val'] > 0].min() / 10 if any(df['adj.P.Val'] > 0) else 1e-10
    df['adj.P.Val'] = df['adj.P.Val'].replace(0, min_pval)
    df['-log10(adj.P.Val)'] = -np.log10(df['adj.P.Val'])

    # Determine significance
    df['significant'] = df['adj.P.Val'] < 0.05

    # Create regulation column
    df['regulation'] = 'not significant'
    df.loc[(df['significant']) & (df['logFC'] > 0), 'regulation'] = 'up-regulated'
    df.loc[(df['significant']) & (df['logFC'] < 0), 'regulation'] = 'down-regulated'

    # Clean data
    df = df.dropna(subset=['EntrezGeneSymbol', 'logFC', 'adj.P.Val'])
    df['logFC'] = pd.to_numeric(df['logFC'], errors='coerce')
    df['adj.P.Val'] = pd.to_numeric(df['adj.P.Val'], errors='coerce')
    df['-log10(adj.P.Val)'] = pd.to_numeric(df['-log10(adj.P.Val)'], errors='coerce')
    df = df.dropna(subset=['logFC', 'adj.P.Val', '-log10(adj.P.Val)'])

    logger.info(f"Loaded volcano data with {len(df)} rows")
    return df


def get_sample_age_group(column_name):
    col_str = str(column_name).upper()
    if 'YD' in col_str:
        return 'Young'
    elif 'OD' in col_str:
        return 'Old'
    return None


@functools.lru_cache(maxsize=50)
def load_boxplot_data(gene_name):
    """Load data for boxplot of a specific gene with caching."""
    file_path = get_data_file_path()
    logger.info(f"Loading boxplot data for gene {gene_name} (first load or cache miss)")

    try:
        # Find the correct header row
        wb = load_workbook(filename=file_path, read_only=True)
        sheet = wb['S4A values']

        header_row = None
        for i in range(1, 10):
            if 'EntrezGeneSymbol' in [str(cell.value).strip() for cell in list(sheet.rows)[i - 1]]:
                header_row = i - 1
                break

        wb.close()

        if header_row is not None:
            df = pd.read_excel(file_path, sheet_name='S4A values', header=header_row)
            logger.info(f"Found header row at position {header_row}")
        else:
            logger.warning("Could not find 'EntrezGeneSymbol' in the first 10 rows. Falling back to header=2.")
            df = pd.read_excel(file_path, sheet_name='S4A values', header=2)

        # Filter by gene name
        gene_data = df[df['EntrezGeneSymbol'] == gene_name]

        if gene_data.empty:
            logger.warning(f"No data found for gene {gene_name} in S4A values sheet")
            return None

        # Get donor columns
        donor_columns = [col for col in df.columns if ('OD' in str(col).upper() or 'YD' in str(col).upper())]

        if not donor_columns:
            logger.warning(f"No donor columns found (containing 'OD' or 'YD')")
            return None

        logger.info(f"Found {len(donor_columns)} donor columns")

        # Create boxplot data
        boxplot_data = []
        for col in donor_columns:
            age_group = get_sample_age_group(col)
            if age_group and not gene_data[col].isnull().all():
                value = gene_data[col].values[0]
                boxplot_data.append({
                    'age_group': age_group,
                    'value': float(value) if not pd.isna(value) else 0.0,
                    'sample': str(col)
                })

        logger.info(f"Created boxplot data with {len(boxplot_data)} points for gene {gene_name}")
        return pd.DataFrame(boxplot_data)

    except Exception as e:
        logger.error(f"Error loading boxplot data for gene {gene_name}: {str(e)}", exc_info=True)
        return None


def get_gene_data(gene_name):
    """Get combined gene data including both volcano plot info and boxplot data."""
    logger.info(f"Getting combined data for gene {gene_name}")
    volcano_data = load_volcano_data()

    # Find gene in volcano data
    gene_volcano_data = volcano_data[volcano_data['EntrezGeneSymbol'] == gene_name]

    if gene_volcano_data.empty:
        logger.warning(f"Gene {gene_name} not found in volcano data")
        return None

    # Get boxplot data
    boxplot_data = load_boxplot_data(gene_name)

    if boxplot_data is None or boxplot_data.empty:
        logger.warning(f"No boxplot data available for gene {gene_name}")
        return None

    # Convert numpy values to native Python types for JSON
    gene_info = {}
    for key, value in gene_volcano_data.iloc[0].to_dict().items():
        if isinstance(value, (np.integer, np.floating)):
            gene_info[key] = float(value)
        elif isinstance(value, np.bool_):
            gene_info[key] = bool(value)
        else:
            gene_info[key] = value

    # Prepare result
    result = {
        'gene_info': gene_info,
        'boxplot_data': boxplot_data.to_dict(orient='records')
    }

    logger.info(f"Successfully compiled data for gene {gene_name}")
    return result